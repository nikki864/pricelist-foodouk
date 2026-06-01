import streamlit as st
import pandas as pd
import io
import json
import os
import datetime
from pathlib import Path

# ─── Page Config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Foodomarket Price List Builder",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Constants ───────────────────────────────────────────────────────────────
SAVED_LISTS_DIR = Path("saved_lists")
SAVED_LISTS_DIR.mkdir(exist_ok=True)
REQUIRED_COLUMNS = {"Product Name", "Supplier", "Category", "Cuisine", "Price", "Unit"}

# ─── CSS ─────────────────────────────────────────────────────────────────────
st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');

/* ── Root Variables ── */
:root {
    --brand-orange:  #F97316;
    --brand-amber:   #FB923C;
    --brand-deep:    #7C2D12;
    --brand-cream:   #FFF7ED;
    --brand-green:   #16A34A;
    --neutral-100:   #F5F5F4;
    --neutral-200:   #E7E5E4;
    --neutral-700:   #44403C;
    --neutral-900:   #1C1917;
}

/* ── Global ── */
html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
    color: var(--neutral-900);
}

/* ── Hide Streamlit chrome ── */
#MainMenu, footer, header {visibility: hidden;}
.block-container {padding-top: 1.5rem; padding-bottom: 3rem;}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: var(--neutral-900) !important;
    border-right: 3px solid var(--brand-orange);
}
section[data-testid="stSidebar"] * {color: #E7E5E4 !important;}
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stMultiSelect label,
section[data-testid="stSidebar"] .stSlider label,
section[data-testid="stSidebar"] .stNumberInput label {
    color: var(--brand-amber) !important;
    font-family: 'Syne', sans-serif;
    font-weight: 600;
    font-size: 0.8rem;
    letter-spacing: 0.06em;
    text-transform: uppercase;
}
section[data-testid="stSidebar"] .stTextInput input,
section[data-testid="stSidebar"] .stSelectbox select {
    background: #292524 !important;
    border: 1px solid #57534E !important;
    color: #E7E5E4 !important;
}
section[data-testid="stSidebar"] [data-baseweb="select"] {
    background: #292524 !important;
}

/* ── Hero Banner ── */
.foodo-hero {
    background: linear-gradient(135deg, var(--neutral-900) 0%, #292524 40%, #3C1B0B 100%);
    border-radius: 16px;
    padding: 2.2rem 2.5rem;
    margin-bottom: 1.8rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border: 1px solid #57534E;
    position: relative;
    overflow: hidden;
}
.foodo-hero::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 220px; height: 220px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(249,115,22,0.18) 0%, transparent 70%);
}
.foodo-hero-left h1 {
    font-family: 'Syne', sans-serif;
    font-weight: 800;
    font-size: 2rem;
    color: #FFFFFF;
    margin: 0 0 0.3rem 0;
    line-height: 1.1;
}
.foodo-hero-left h1 span {color: var(--brand-orange);}
.foodo-hero-left p {
    color: #A8A29E;
    font-size: 0.92rem;
    margin: 0;
}
.foodo-logo-badge {
    background: var(--brand-orange);
    color: white;
    font-family: 'Syne', sans-serif;
    font-weight: 800;
    font-size: 1.05rem;
    padding: 0.55rem 1.1rem;
    border-radius: 8px;
    letter-spacing: 0.02em;
    white-space: nowrap;
}

/* ── Section Headers ── */
.section-title {
    font-family: 'Syne', sans-serif;
    font-weight: 700;
    font-size: 1rem;
    color: var(--neutral-900);
    text-transform: uppercase;
    letter-spacing: 0.07em;
    border-left: 4px solid var(--brand-orange);
    padding-left: 0.75rem;
    margin: 1.6rem 0 1rem 0;
}

/* ── Cards ── */
.metric-card {
    background: white;
    border: 1px solid var(--neutral-200);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    border-top: 3px solid var(--brand-orange);
}
.metric-card .metric-val {
    font-family: 'Syne', sans-serif;
    font-weight: 800;
    font-size: 1.8rem;
    color: var(--brand-orange);
    line-height: 1;
}
.metric-card .metric-label {
    font-size: 0.78rem;
    color: #78716C;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-top: 0.25rem;
}

/* ── Buttons ── */
.stButton > button {
    background: var(--brand-orange) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    letter-spacing: 0.03em !important;
    padding: 0.5rem 1.4rem !important;
    transition: background 0.2s !important;
}
.stButton > button:hover {
    background: var(--brand-deep) !important;
}

/* ── Download buttons ── */
.stDownloadButton > button {
    background: var(--neutral-900) !important;
    color: white !important;
    border: 1px solid var(--brand-orange) !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
}
.stDownloadButton > button:hover {
    background: var(--brand-orange) !important;
}

/* ── Success / Info alerts ── */
.stSuccess, .stInfo {border-radius: 10px !important;}

/* ── Dataframe ── */
.stDataFrame {border: 1px solid var(--neutral-200) !important; border-radius: 8px !important;}

/* ── Expander ── */
.stExpander {border: 1px solid var(--neutral-200) !important; border-radius: 8px !important;}

/* ── Tabs ── */
.stTabs [data-baseweb="tab"] {
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
}
.stTabs [aria-selected="true"] {
    color: var(--brand-orange) !important;
    border-bottom-color: var(--brand-orange) !important;
}

/* ── Upload area ── */
.uploadedFile {border-color: var(--brand-orange) !important;}

/* ── Final price list table ── */
.price-list-header {
    background: var(--neutral-900);
    color: white;
    font-family: 'Syne', sans-serif;
    font-weight: 700;
    padding: 0.9rem 1.2rem;
    border-radius: 10px 10px 0 0;
    display: flex;
    gap: 2rem;
    font-size: 0.82rem;
    letter-spacing: 0.06em;
    text-transform: uppercase;
}
.saved-list-card {
    background: white;
    border: 1px solid var(--neutral-200);
    border-left: 4px solid var(--brand-orange);
    border-radius: 8px;
    padding: 0.9rem 1.1rem;
    margin-bottom: 0.6rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.saved-list-name {
    font-family: 'Syne', sans-serif;
    font-weight: 700;
    font-size: 0.95rem;
}
.saved-list-meta {
    font-size: 0.78rem;
    color: #78716C;
    margin-top: 0.1rem;
}
</style>
""",
    unsafe_allow_html=True,
)

# ─── Hero ─────────────────────────────────────────────────────────────────────
st.markdown(
    """
<div class="foodo-hero">
  <div class="foodo-hero-left">
    <h1>Price List <span>Builder</span></h1>
    <p>Upload your product database · Filter · Curate · Export</p>
  </div>
  <div class="foodo-logo-badge">🛒 FOODOMARKET</div>
</div>
""",
    unsafe_allow_html=True,
)

# ─── Session State Init ───────────────────────────────────────────────────────
for key, val in {
    "df": None,
    "selected_ids": set(),
    "custom_products": [],
    "final_list": None,
}.items():
    if key not in st.session_state:
        st.session_state[key] = val


# ─── Helpers ─────────────────────────────────────────────────────────────────

def load_excel(file) -> pd.DataFrame | None:
    try:
        df = pd.read_excel(file, engine="openpyxl")
        df.columns = df.columns.str.strip()
        missing = REQUIRED_COLUMNS - set(df.columns)
        if missing:
            st.error(f"Missing columns: {', '.join(missing)}")
            return None
        df["_id"] = range(len(df))
        df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0)
        return df
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return None


def export_excel(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"

    # Header row
    headers = ["#", "Product", "Supplier", "Price", "Unit"]
    header_fill = PatternFill("solid", start_color="1C1917")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 24

    thin = Side(style="thin", color="E7E5E4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    orange_fill = PatternFill("solid", start_color="FFF7ED")

    for i, (_, row) in enumerate(df.iterrows(), 1):
        fill = orange_fill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        vals = [i, row.get("Product", row.get("Product Name", "")),
                row.get("Supplier", ""), row.get("Price", 0), row.get("Unit", "")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if col == 4:  # Price
                cell.number_format = '#,##0.00 "TND"'
                cell.alignment = Alignment(horizontal="right")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    # Total row
    total_row = len(df) + 2
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{len(df)+1})").number_format = '#,##0.00 "TND"'
    ws.cell(row=total_row, column=4).font = Font(bold=True, name="Arial", color="F97316")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(df: pd.DataFrame, list_name: str = "Price List") -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=18*mm, bottomMargin=18*mm,
                            leftMargin=18*mm, rightMargin=18*mm)

    styles = getSampleStyleSheet()
    orange = colors.HexColor("#F97316")
    dark   = colors.HexColor("#1C1917")
    cream  = colors.HexColor("#FFF7ED")

    title_style = ParagraphStyle("Title", fontName="Helvetica-Bold",
                                 fontSize=20, textColor=colors.white,
                                 spaceAfter=2, alignment=TA_CENTER)
    sub_style   = ParagraphStyle("Sub", fontName="Helvetica",
                                 fontSize=9, textColor=colors.HexColor("#A8A29E"),
                                 alignment=TA_CENTER)

    elems = []

    # Header block
    header_data = [[Paragraph(f"🛒  FOODOMARKET", title_style)],
                   [Paragraph(list_name, sub_style)],
                   [Paragraph(f"Generated: {datetime.datetime.now().strftime('%d %B %Y  %H:%M')}", sub_style)]]
    header_tbl = Table(header_data, colWidths=[174*mm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), dark),
        ("TOPPADDING",    (0,0), (-1,0), 14),
        ("BOTTOMPADDING", (0,-1), (-1,-1), 12),
        ("ROUNDEDCORNERS", [8]),
    ]))
    elems.append(header_tbl)
    elems.append(Spacer(1, 10*mm))

    # Table
    col_headers = ["#", "Product", "Supplier", "Price", "Unit"]
    rows = [col_headers]
    total = 0
    for i, (_, row) in enumerate(df.iterrows(), 1):
        price = float(row.get("Price", 0))
        total += price
        rows.append([
            str(i),
            str(row.get("Product", row.get("Product Name", ""))),
            str(row.get("Supplier", "")),
            f"{price:.3f} TND",
            str(row.get("Unit", "")),
        ])
    rows.append(["", "", "TOTAL", f"{total:.3f} TND", ""])

    col_widths = [10*mm, 74*mm, 44*mm, 28*mm, 18*mm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        # Header
        ("BACKGROUND",  (0,0), (-1,0), dark),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 8.5),
        ("TOPPADDING",  (0,0), (-1,0), 8),
        ("BOTTOMPADDING",(0,0),(-1,0), 8),
        ("ALIGN",       (0,0), (-1,0), "CENTER"),
        # Body
        ("FONTNAME",    (0,1), (-1,-2), "Helvetica"),
        ("FONTSIZE",    (0,1), (-1,-2), 8.5),
        ("TOPPADDING",  (0,1), (-1,-2), 5),
        ("BOTTOMPADDING",(0,1),(-1,-2), 5),
        ("ALIGN",       (3,1), (3,-1), "RIGHT"),
        ("ALIGN",       (0,1), (0,-1), "CENTER"),
        # Alternating
        *[("BACKGROUND", (0, r), (-1, r), cream)
          for r in range(2, len(rows)-1, 2)],
        # Grid
        ("LINEBELOW",   (0,0), (-1,0), 0.5, orange),
        ("LINEBELOW",   (0,1), (-1,-2), 0.3, colors.HexColor("#E7E5E4")),
        # Total row
        ("FONTNAME",    (0,-1), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND",  (0,-1), (-1,-1), colors.HexColor("#FFF7ED")),
        ("LINEABOVE",   (0,-1), (-1,-1), 1, orange),
        ("TEXTCOLOR",   (3,-1), (3,-1), orange),
    ])
    tbl.setStyle(style)
    elems.append(tbl)

    # Footer
    elems.append(Spacer(1, 8*mm))
    footer = Paragraph(
        f"<font color='#78716C' size='7'>© {datetime.datetime.now().year} Foodomarket · "
        f"Confidential Price List · {len(df)} items</font>",
        ParagraphStyle("footer", alignment=TA_CENTER)
    )
    elems.append(footer)

    doc.build(elems)
    return buf.getvalue()


def save_list(name: str, df: pd.DataFrame):
    path = SAVED_LISTS_DIR / f"{name}.json"
    data = {
        "name": name,
        "created": datetime.datetime.now().isoformat(),
        "items": df.to_dict(orient="records"),
    }
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2))


def load_saved_lists() -> list[dict]:
    lists = []
    for path in sorted(SAVED_LISTS_DIR.glob("*.json"), reverse=True):
        try:
            lists.append(json.loads(path.read_text()))
        except Exception:
            pass
    return lists


def delete_saved_list(name: str):
    path = SAVED_LISTS_DIR / f"{name}.json"
    if path.exists():
        path.unlink()


# ─── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        "<div style='font-family:Syne,sans-serif;font-size:1.15rem;"
        "font-weight:800;color:#F97316;padding-bottom:0.8rem;"
        "border-bottom:1px solid #44403C;margin-bottom:1rem;'>"
        "⚙️ Filters & Settings</div>",
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader("Upload Product Database (.xlsx)", type=["xlsx"])
    if uploaded:
        df_loaded = load_excel(uploaded)
        if df_loaded is not None:
            st.session_state.df = df_loaded
            st.success(f"✅ {len(df_loaded)} products loaded")

    st.divider()

    df_main: pd.DataFrame | None = st.session_state.df

    if df_main is not None:
        cuisines = ["All"] + sorted(df_main["Cuisine"].dropna().unique().tolist())
        sel_cuisine = st.selectbox("Cuisine", cuisines)

        suppliers = sorted(df_main["Supplier"].dropna().unique().tolist())
        sel_suppliers = st.multiselect("Supplier", suppliers, default=[])

        categories = sorted(df_main["Category"].dropna().unique().tolist())
        sel_categories = st.multiselect("Category", categories, default=[])

        st.divider()

        max_products = len(df_main)
        num_products = st.number_input(
            "Max products to show",
            min_value=1,
            max_value=max(max_products, 1),
            value=min(50, max_products),
            step=5,
        )
    else:
        st.info("Upload a file to enable filters.")
        sel_cuisine = "All"
        sel_suppliers = []
        sel_categories = []
        num_products = 50

# ─── Main Tabs ────────────────────────────────────────────────────────────────
tab_build, tab_generate, tab_saved = st.tabs(
    ["📋  Build List", "✅  Generate & Export", "📁  Saved Lists"]
)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — BUILD
# ═══════════════════════════════════════════════════════════════════════════════
with tab_build:

    if st.session_state.df is None:
        st.info("👆 Upload a product database Excel file using the sidebar to get started.")
        st.stop()

    df_main: pd.DataFrame = st.session_state.df

    # ── Apply Filters ──────────────────────────────────────────────────────
    filtered = df_main.copy()
    if sel_cuisine != "All":
        filtered = filtered[filtered["Cuisine"] == sel_cuisine]
    if sel_suppliers:
        filtered = filtered[filtered["Supplier"].isin(sel_suppliers)]
    if sel_categories:
        filtered = filtered[filtered["Category"].isin(sel_categories)]

    # ── Search ────────────────────────────────────────────────────────────
    st.markdown('<div class="section-title">🔍 Search Products</div>', unsafe_allow_html=True)
    search_q = st.text_input("Search by product name", placeholder="e.g. olive oil, chicken…", label_visibility="collapsed")
    if search_q.strip():
        filtered = filtered[filtered["Product Name"].str.contains(search_q.strip(), case=False, na=False)]

    filtered = filtered.head(num_products)

    # ── Metrics ───────────────────────────────────────────────────────────
    st.markdown('<div class="section-title">📊 Overview</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="metric-card"><div class="metric-val">{len(filtered)}</div>'
                    f'<div class="metric-label">Matching Products</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="metric-card"><div class="metric-val">{len(st.session_state.selected_ids)}</div>'
                    f'<div class="metric-label">Selected</div></div>', unsafe_allow_html=True)
    with c3:
        sel_total = sum(
            float(df_main.loc[df_main["_id"] == i, "Price"].values[0])
            for i in st.session_state.selected_ids
            if not df_main.loc[df_main["_id"] == i].empty
        )
        custom_total = sum(float(p.get("Price", 0)) for p in st.session_state.custom_products)
        grand_total = sel_total + custom_total
        st.markdown(f'<div class="metric-card"><div class="metric-val">{grand_total:.2f}</div>'
                    f'<div class="metric-label">Est. Total (TND)</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="metric-card"><div class="metric-val">{len(st.session_state.custom_products)}</div>'
                    f'<div class="metric-label">Custom Products</div></div>', unsafe_allow_html=True)

    # ── Product Table ──────────────────────────────────────────────────────
    st.markdown('<div class="section-title">🧺 Product Selection</div>', unsafe_allow_html=True)

    if len(filtered) == 0:
        st.warning("No products match your filters.")
    else:
        col_chk, col_tbl = st.columns([0.05, 0.95])

        # Select / Deselect All
        btn_col1, btn_col2, _ = st.columns([1, 1, 6])
        with btn_col1:
            if st.button("Select All"):
                for pid in filtered["_id"].tolist():
                    st.session_state.selected_ids.add(pid)
                st.rerun()
        with btn_col2:
            if st.button("Deselect All"):
                for pid in filtered["_id"].tolist():
                    st.session_state.selected_ids.discard(pid)
                st.rerun()

        # Render product rows
        for _, row in filtered.iterrows():
            pid = row["_id"]
            checked = pid in st.session_state.selected_ids
            r1, r2, r3, r4, r5, r6 = st.columns([0.5, 3.5, 2, 1.5, 1.2, 1.2])
            new_check = r1.checkbox("", value=checked, key=f"chk_{pid}", label_visibility="collapsed")
            if new_check != checked:
                if new_check:
                    st.session_state.selected_ids.add(pid)
                else:
                    st.session_state.selected_ids.discard(pid)
                st.rerun()
            r2.markdown(f"**{row['Product Name']}**")
            r3.markdown(f"<span style='color:#78716C'>{row['Supplier']}</span>", unsafe_allow_html=True)
            r4.markdown(f"<span style='color:#16A34A;font-weight:600'>{row['Price']:.3f} TND</span>", unsafe_allow_html=True)
            r5.markdown(f"<span style='color:#78716C'>{row['Unit']}</span>", unsafe_allow_html=True)
            r6.markdown(
                f"<span style='background:#F3F4F6;color:#6B7280;font-size:0.72rem;"
                f"padding:2px 7px;border-radius:12px'>{row['Category']}</span>",
                unsafe_allow_html=True,
            )

    # ── Manual Additions ───────────────────────────────────────────────────
    st.markdown('<div class="section-title">➕ Add Custom Product</div>', unsafe_allow_html=True)
    with st.expander("Add a product not in the database"):
        a1, a2, a3, a4 = st.columns([3, 2, 1.2, 1])
        c_name     = a1.text_input("Product Name", key="c_name", placeholder="e.g. Special Blend Tea")
        c_supplier = a2.text_input("Supplier",     key="c_supplier", placeholder="Supplier name")
        c_price    = a3.number_input("Price (TND)", key="c_price", min_value=0.0, step=0.5, format="%.3f")
        c_unit     = a4.text_input("Unit",         key="c_unit", placeholder="kg / L…")

        if st.button("Add to List"):
            if c_name.strip():
                st.session_state.custom_products.append({
                    "Product": c_name.strip(),
                    "Supplier": c_supplier.strip() or "—",
                    "Price": c_price,
                    "Unit": c_unit.strip() or "—",
                })
                st.success(f"Added **{c_name.strip()}**")
                st.rerun()
            else:
                st.warning("Product name is required.")

    if st.session_state.custom_products:
        st.markdown("**Custom products added:**")
        for idx, cp in enumerate(st.session_state.custom_products):
            ca, cb, cc = st.columns([4, 2, 1])
            ca.markdown(f"• **{cp['Product']}** — {cp['Supplier']}")
            cb.markdown(f"<span style='color:#16A34A;font-weight:600'>{float(cp['Price']):.3f} TND / {cp['Unit']}</span>", unsafe_allow_html=True)
            if cc.button("🗑", key=f"del_custom_{idx}"):
                st.session_state.custom_products.pop(idx)
                st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — GENERATE & EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
with tab_generate:

    if st.session_state.df is None:
        st.info("Upload a database first.")
        st.stop()

    df_main: pd.DataFrame = st.session_state.df
    n_sel = len(st.session_state.selected_ids)
    n_cus = len(st.session_state.custom_products)

    st.markdown('<div class="section-title">📄 Generate Price List</div>', unsafe_allow_html=True)

    list_name = st.text_input("Price list name", value=f"Price List – {datetime.date.today().strftime('%d %b %Y')}")

    col_gen1, col_gen2 = st.columns([1, 3])
    with col_gen1:
        generate_btn = st.button("🚀 Generate Price List", use_container_width=True)

    if generate_btn:
        rows = []
        for pid in sorted(st.session_state.selected_ids):
            sub = df_main[df_main["_id"] == pid]
            if not sub.empty:
                r = sub.iloc[0]
                rows.append({
                    "Product": r["Product Name"],
                    "Supplier": r["Supplier"],
                    "Price": float(r["Price"]),
                    "Unit": r["Unit"],
                })
        for cp in st.session_state.custom_products:
            rows.append({
                "Product":  cp["Product"],
                "Supplier": cp["Supplier"],
                "Price":    float(cp["Price"]),
                "Unit":     cp["Unit"],
            })

        if not rows:
            st.warning("No products selected. Go to **Build List** and check some products.")
        else:
            st.session_state.final_list = pd.DataFrame(rows)
            st.session_state.final_list_name = list_name
            st.success(f"✅ Price list generated with **{len(rows)} products**")

    if st.session_state.final_list is not None:
        fl = st.session_state.final_list
        ln = st.session_state.get("final_list_name", "Price List")

        st.markdown(f'<div class="section-title">📋 {ln}</div>', unsafe_allow_html=True)

        # Display table
        display_fl = fl.copy()
        display_fl.index = range(1, len(display_fl) + 1)
        display_fl["Price"] = display_fl["Price"].map(lambda x: f"{x:.3f} TND")
        st.dataframe(display_fl, use_container_width=True, height=min(400, 40 + 36 * len(fl)))

        total = fl["Price"].sum()
        st.markdown(
            f"<div style='text-align:right;font-family:Syne,sans-serif;"
            f"font-size:1.1rem;font-weight:700;color:#F97316;margin-top:0.3rem'>"
            f"Total: {total:.3f} TND</div>",
            unsafe_allow_html=True,
        )

        # Export buttons
        st.markdown('<div class="section-title">📤 Export</div>', unsafe_allow_html=True)
        e1, e2, e3 = st.columns(3)

        with e1:
            xlsx_bytes = export_excel(fl)
            st.download_button(
                "⬇️ Download Excel (.xlsx)",
                data=xlsx_bytes,
                file_name=f"{ln.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with e2:
            pdf_bytes = export_pdf(fl, ln)
            st.download_button(
                "⬇️ Download PDF (.pdf)",
                data=pdf_bytes,
                file_name=f"{ln.replace(' ', '_')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        with e3:
            if st.button("💾 Save for Future Reuse", use_container_width=True):
                save_list(ln, fl)
                st.success(f'Saved as "{ln}"')


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — SAVED LISTS
# ═══════════════════════════════════════════════════════════════════════════════
with tab_saved:
    st.markdown('<div class="section-title">📁 Saved Price Lists</div>', unsafe_allow_html=True)
    saved = load_saved_lists()

    if not saved:
        st.info("No saved lists yet. Generate a price list and click **Save for Future Reuse**.")
    else:
        for lst in saved:
            created_str = lst.get("created", "")[:16].replace("T", " ")
            item_count  = len(lst.get("items", []))

            st.markdown(
                f'<div class="saved-list-card">'
                f'<div><div class="saved-list-name">📋 {lst["name"]}</div>'
                f'<div class="saved-list-meta">{created_str} · {item_count} items</div></div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            df_saved = pd.DataFrame(lst.get("items", []))
            if not df_saved.empty:
                s1, s2, s3, s4 = st.columns([2, 2, 2, 1])

                with s1:
                    with st.expander("👁 Preview"):
                        preview = df_saved.copy()
                        if "Price" in preview.columns:
                            preview["Price"] = preview["Price"].map(lambda x: f"{float(x):.3f} TND")
                        st.dataframe(preview, use_container_width=True)

                with s2:
                    xlsx_b = export_excel(df_saved)
                    st.download_button(
                        "⬇️ Excel",
                        data=xlsx_b,
                        file_name=f"{lst['name'].replace(' ','_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_xl_{lst['name']}",
                        use_container_width=True,
                    )

                with s3:
                    pdf_b = export_pdf(df_saved, lst["name"])
                    st.download_button(
                        "⬇️ PDF",
                        data=pdf_b,
                        file_name=f"{lst['name'].replace(' ','_')}.pdf",
                        mime="application/pdf",
                        key=f"dl_pdf_{lst['name']}",
                        use_container_width=True,
                    )

                with s4:
                    if st.button("🗑 Delete", key=f"del_{lst['name']}"):
                        delete_saved_list(lst["name"])
                        st.rerun()

            st.divider()
