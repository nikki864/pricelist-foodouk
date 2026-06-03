"""
Foodomarket AI Price List Builder
- Reads any Excel file format
- AI detects: product name, conditioning/unit, price columns
- AI auto-builds cuisine-based lists of 10 products
- Supports 1 or 2 uploaded files
- Category filtering + manual overrides
"""

import streamlit as st
import pandas as pd
import io
import json
import re
import anthropic
import datetime
from pathlib import Path

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Foodomarket AI List Builder",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Styling ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;700;800&family=DM+Sans:wght@300;400;500&display=swap');

:root {
    --or: #F97316; --dk: #1C1917; --cr: #FFF7ED;
    --gr: #16A34A; --sl: #44403C; --lt: #F5F5F4;
}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;color:var(--dk);}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding-top:1.2rem;padding-bottom:2rem;}

/* Sidebar */
section[data-testid="stSidebar"]{background:var(--dk)!important;border-right:3px solid var(--or);}
section[data-testid="stSidebar"] *{color:#E7E5E4!important;}
section[data-testid="stSidebar"] label{
    color:var(--or)!important;font-family:'Syne',sans-serif;font-weight:700;
    font-size:.75rem;letter-spacing:.07em;text-transform:uppercase;}

/* Hero */
.hero{background:linear-gradient(135deg,#1C1917 0%,#292524 50%,#3C1B0B 100%);
    border-radius:14px;padding:1.8rem 2rem;margin-bottom:1.5rem;
    display:flex;justify-content:space-between;align-items:center;
    border:1px solid #44403C;}
.hero h1{font-family:'Syne',sans-serif;font-weight:800;font-size:1.9rem;
    color:#fff;margin:0 0 .2rem 0;}
.hero h1 span{color:var(--or);}
.hero p{color:#A8A29E;font-size:.88rem;margin:0;}
.badge{background:var(--or);color:#fff;font-family:'Syne',sans-serif;font-weight:800;
    font-size:.95rem;padding:.5rem 1rem;border-radius:8px;}

/* Section titles */
.st{font-family:'Syne',sans-serif;font-weight:700;font-size:.9rem;color:var(--dk);
    text-transform:uppercase;letter-spacing:.07em;border-left:4px solid var(--or);
    padding-left:.7rem;margin:1.4rem 0 .8rem 0;}

/* Buttons */
.stButton>button{background:var(--or)!important;color:#fff!important;border:none!important;
    border-radius:8px!important;font-family:'Syne',sans-serif!important;font-weight:700!important;}
.stButton>button:hover{background:#C2410C!important;}
.stDownloadButton>button{background:var(--dk)!important;color:#fff!important;
    border:1px solid var(--or)!important;border-radius:8px!important;
    font-family:'Syne',sans-serif!important;font-weight:600!important;}
.stDownloadButton>button:hover{background:var(--or)!important;}

/* Cards */
.card{background:#fff;border:1px solid #E7E5E4;border-radius:10px;
    padding:.9rem 1.1rem;border-top:3px solid var(--or);margin-bottom:.6rem;}
.card .val{font-family:'Syne',sans-serif;font-weight:800;font-size:1.6rem;color:var(--or);}
.card .lbl{font-size:.73rem;color:#78716C;text-transform:uppercase;letter-spacing:.05em;}

/* Product row */
.prow{background:#fff;border:1px solid #E7E5E4;border-radius:8px;
    padding:.6rem 1rem;margin:.3rem 0;display:flex;align-items:center;gap:.8rem;}
.prow:hover{border-color:var(--or);background:var(--cr);}
.pname{font-weight:600;font-size:.9rem;flex:2;}
.psup{color:#78716C;font-size:.82rem;flex:1.5;}
.pprice{color:var(--gr);font-weight:700;font-size:.9rem;flex:.8;}
.punit{color:#78716C;font-size:.8rem;flex:.7;}
.pcat{background:#F3F4F6;color:#6B7280;font-size:.7rem;padding:2px 8px;
    border-radius:12px;white-space:nowrap;}

/* AI thinking */
.ai-box{background:linear-gradient(135deg,#1C1917,#292524);border-radius:10px;
    padding:1.2rem 1.5rem;border-left:4px solid var(--or);color:#E7E5E4;
    font-family:'DM Sans',sans-serif;font-size:.9rem;line-height:1.6;margin:.8rem 0;}
.ai-box .ai-title{font-family:'Syne',sans-serif;color:var(--or);font-weight:700;
    font-size:.8rem;letter-spacing:.06em;text-transform:uppercase;margin-bottom:.5rem;}

/* Chip */
.chip{display:inline-block;background:var(--or);color:#fff;font-size:.72rem;
    font-weight:700;padding:2px 9px;border-radius:12px;margin:1px;font-family:'Syne',sans-serif;}
.chip-g{background:var(--gr);}
</style>
""", unsafe_allow_html=True)

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div>
    <h1>AI Price List <span>Builder</span></h1>
    <p>Upload any Excel · AI detects columns · Choose cuisine · Auto-build lists of 10</p>
  </div>
  <div class="badge">🤖 FOODOMARKET AI</div>
</div>
""", unsafe_allow_html=True)

# ── Session state ─────────────────────────────────────────────────────────────
defaults = dict(
    dfs=[],              # list of DataFrames (1 or 2 files)
    file_names=[],
    mappings=[],         # AI-detected column mappings per file
    unified=None,        # merged & normalised DataFrame
    ai_analysis="",      # AI explanation of column detection
    generated_list=None, # final list of 10 products
    ai_reasoning="",     # AI reasoning for product selection
    detected_categories=[],
    detected_cuisines=[],
)
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

client = anthropic.Anthropic()

# ── Helpers ───────────────────────────────────────────────────────────────────

def read_any_excel(file) -> pd.DataFrame | None:
    """Try several strategies to read any Excel file."""
    try:
        # Try default first
        df = pd.read_excel(file, engine="openpyxl")
        if len(df.columns) >= 2:
            return df
    except Exception:
        pass
    # Try with header on row 1, 2, 3
    for skip in [1, 2, 3]:
        try:
            file.seek(0)
            df = pd.read_excel(file, engine="openpyxl", header=skip)
            if len(df.columns) >= 2:
                return df
        except Exception:
            pass
    # Try reading all sheets, pick the largest
    try:
        file.seek(0)
        sheets = pd.read_excel(file, engine="openpyxl", sheet_name=None)
        best = max(sheets.values(), key=lambda d: len(d))
        return best
    except Exception:
        pass
    return None


def ai_detect_columns(df: pd.DataFrame, filename: str) -> dict:
    """Ask Claude to identify name/conditioning/price columns."""
    sample = df.head(8).to_string(max_cols=20)
    cols = list(df.columns)

    prompt = f"""
You are analyzing an Excel product database for a food distributor.
File: {filename}
Columns available: {cols}
First 8 rows sample:
{sample}

Your job: identify which columns contain:
1. product_name  – the product/article name
2. conditioning  – unit, packaging, weight, volume, format (e.g. 1kg, 6x1L, carton, piece)
3. price         – the selling price (numeric). If multiple price columns exist pick the most relevant retail/unit price.
4. category      – product category or family (may not exist)
5. supplier      – supplier or brand (may not exist)
6. cuisine       – cuisine type (may not exist, often absent)

Return ONLY a valid JSON object like:
{{
  "product_name": "exact column name or null",
  "conditioning": "exact column name or null",
  "price": "exact column name or null",
  "category": "exact column name or null",
  "supplier": "exact column name or null",
  "cuisine": "exact column name or null",
  "explanation": "1-2 sentences describing what you found and any assumptions"
}}
"""
    resp = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=600,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip()
    # Strip markdown fences if present
    raw = re.sub(r"```(?:json)?|```", "", raw).strip()
    return json.loads(raw)


def normalise_df(df: pd.DataFrame, mapping: dict, source_name: str) -> pd.DataFrame:
    """Rename columns to standard names using AI mapping."""
    rename = {}
    for standard, col in mapping.items():
        if col and col in df.columns and standard != "explanation":
            rename[col] = standard
    out = df.rename(columns=rename)
    needed = ["product_name", "conditioning", "price", "category", "supplier", "cuisine"]
    for c in needed:
        if c not in out.columns:
            out[c] = None
    out["_source"] = source_name
    out["price"] = pd.to_numeric(out["price"], errors="coerce")
    out = out.dropna(subset=["product_name"])
    out = out[out["product_name"].astype(str).str.strip() != ""]
    return out[["product_name", "conditioning", "price", "category", "supplier", "cuisine", "_source"]]


def ai_infer_cuisines_and_categories(df: pd.DataFrame) -> tuple[list, list]:
    """Ask AI to extract/infer cuisine types and categories from product names."""
    sample_products = df["product_name"].dropna().astype(str).head(60).tolist()
    cats_raw = df["category"].dropna().astype(str).unique().tolist()[:40]

    prompt = f"""
You are a food distribution expert. Given these product names from a supplier catalog:
{sample_products}

And these raw category values (may be empty or in French/Arabic/other):
{cats_raw}

1. Infer the cuisine types these products belong to from: 
   Moroccan, Tunisian, Algerian, Lebanese, Turkish, Italian, French, Asian, American, Mediterranean, Middle Eastern, International, Other
   Only include cuisines clearly represented by multiple products.

2. List the distinct product categories present (clean, normalised English names).

Return ONLY valid JSON:
{{
  "cuisines": ["Tunisian", "Italian", ...],
  "categories": ["Spices & Herbs", "Grains & Rice", ...]
}}
"""
    resp = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = re.sub(r"```(?:json)?|```", "", resp.content[0].text).strip()
    data = json.loads(raw)
    return data.get("cuisines", []), data.get("categories", [])


def ai_build_list(
    df: pd.DataFrame,
    cuisine: str,
    categories: list[str],
    n: int = 10,
    extra_instruction: str = "",
) -> tuple[list[dict], str]:
    """Ask AI to select the best N products for a cuisine from the dataframe."""
    # Prepare a compact product list for the AI
    rows = []
    for i, (_, r) in enumerate(df.iterrows()):
        rows.append({
            "id": i,
            "name": str(r.get("product_name", "")),
            "cond": str(r.get("conditioning", "") or ""),
            "price": r.get("price", None),
            "cat": str(r.get("category", "") or ""),
            "supplier": str(r.get("supplier", "") or ""),
            "source": str(r.get("_source", "")),
        })

    cat_filter = f"Focus on these categories: {categories}." if categories else ""
    extra = f"Additional instruction: {extra_instruction}." if extra_instruction else ""

    prompt = f"""
You are an expert food distributor buyer for a {cuisine} cuisine restaurant/caterer.

Select exactly {n} products from this catalog that are most relevant for {cuisine} cuisine cooking.
{cat_filter}
{extra}

Rules:
- Pick products that a {cuisine} cuisine kitchen would actually use
- Prefer variety across categories (don't pick 5 oils)
- Prefer products with a valid numeric price
- If products come from 2 sources, try to balance (4-6 from each)
- Return a diverse, practical selection

Product catalog (JSON):
{json.dumps(rows, ensure_ascii=False)}

Return ONLY valid JSON:
{{
  "selected_ids": [3, 17, 42, ...],
  "reasoning": "2-3 sentences explaining your selection logic"
}}
"""
    resp = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=800,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = re.sub(r"```(?:json)?|```", "", resp.content[0].text).strip()
    data = json.loads(raw)
    selected_ids = data.get("selected_ids", [])
    reasoning = data.get("reasoning", "")

    selected_rows = [rows[i] for i in selected_ids if i < len(rows)]
    return selected_rows, reasoning


def export_excel(items: list[dict], title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"
    headers = ["#", "Product", "Conditioning", "Price", "Category", "Supplier", "Source"]
    dark = "1C1917"; orange = "F97316"; cream = "FFF7ED"

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", start_color=dark)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for i, item in enumerate(items, 1):
        fill = PatternFill("solid", start_color=cream if i%2==0 else "FFFFFF")
        vals = [i, item["name"], item["cond"], item.get("price"), item["cat"], item["supplier"], item["source"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i+1, column=c, value=v)
            cell.fill = fill
            cell.font = Font(name="Arial", size=9)
            if c == 4 and v is not None:
                try:
                    cell.value = float(v)
                    cell.number_format = '#,##0.000 "TND"'
                except Exception:
                    pass
            if c == 1:
                cell.alignment = Alignment(horizontal="center")

    for col, width in zip("ABCDEFG", [5, 38, 16, 14, 20, 22, 14]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(items: list[dict], title: str, cuisine: str) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    orange = colors.HexColor("#F97316")
    dark   = colors.HexColor("#1C1917")
    cream  = colors.HexColor("#FFF7ED")

    ts = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=18, textColor=colors.white,
                        alignment=TA_CENTER, spaceAfter=2)
    ss = ParagraphStyle("s", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#A8A29E"),
                        alignment=TA_CENTER)
    elems = []

    hdr = Table([[Paragraph(f"🛒 FOODOMARKET – {cuisine.upper()} PRICE LIST", ts)],
                 [Paragraph(f"{title} · {datetime.date.today().strftime('%d %B %Y')}", ss)]],
                colWidths=[180*mm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), dark),
        ("TOPPADDING", (0,0),(-1,0), 14),
        ("BOTTOMPADDING",(0,-1),(-1,-1),10),
    ]))
    elems += [hdr, Spacer(1, 8*mm)]

    rows = [["#", "Product", "Conditioning", "Price", "Category", "Source"]]
    total = 0
    for i, item in enumerate(items, 1):
        p = item.get("price")
        try:
            p_fmt = f"{float(p):.3f} TND"
            total += float(p)
        except Exception:
            p_fmt = "—"
        rows.append([str(i), item["name"], item["cond"], p_fmt, item["cat"], item["source"]])
    rows.append(["", "", "", f"{total:.3f} TND", "TOTAL", ""])

    tbl = Table(rows, colWidths=[8*mm,68*mm,28*mm,26*mm,32*mm,18*mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), dark), ("TEXTCOLOR",(0,0),(-1,0), colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,0),8),
        ("TOPPADDING",(0,0),(-1,0),7), ("BOTTOMPADDING",(0,0),(-1,0),7),
        ("ALIGN",(0,0),(-1,0),"CENTER"),
        ("FONTNAME",(0,1),(-1,-2),"Helvetica"), ("FONTSIZE",(0,1),(-1,-2),8),
        ("TOPPADDING",(0,1),(-1,-2),5), ("BOTTOMPADDING",(0,1),(-1,-2),5),
        ("ALIGN",(3,1),(3,-1),"RIGHT"), ("ALIGN",(0,1),(0,-1),"CENTER"),
        *[("BACKGROUND",(0,r),(-1,r),cream) for r in range(2,len(rows)-1,2)],
        ("LINEBELOW",(0,0),(-1,0),.5,orange),
        ("LINEBELOW",(0,1),(-1,-2),.3,colors.HexColor("#E7E5E4")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("BACKGROUND",(0,-1),(-1,-1),cream),
        ("LINEABOVE",(0,-1),(-1,-1),1,orange),
        ("TEXTCOLOR",(3,-1),(3,-1),orange),
    ]))
    elems.append(tbl)
    doc.build(elems)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(
        "<div style='font-family:Syne,sans-serif;font-size:1rem;font-weight:800;"
        "color:#F97316;padding-bottom:.7rem;border-bottom:1px solid #44403C;"
        "margin-bottom:1rem'>⚙️ Upload & Settings</div>",
        unsafe_allow_html=True,
    )

    f1 = st.file_uploader("Excel File 1 (required)", type=["xlsx","xls"], key="fu1")
    f2 = st.file_uploader("Excel File 2 (optional)", type=["xlsx","xls"], key="fu2")

    st.divider()

    analyze_btn = st.button("🔍 Analyse Files with AI", use_container_width=True)
    st.divider()

    # Cuisine picker (populated after analysis)
    cuisine_options = st.session_state.detected_cuisines or [
        "Tunisian","Moroccan","Algerian","Lebanese","Turkish",
        "Italian","French","Asian","Mediterranean","International"
    ]
    selected_cuisine = st.selectbox("🍽️ Target Cuisine", cuisine_options)

    # Category filter
    all_cats = st.session_state.detected_categories
    if all_cats:
        selected_cats = st.multiselect("📦 Filter Categories (optional)", all_cats)
    else:
        selected_cats = []

    n_products = st.number_input("Number of products", min_value=5, max_value=30, value=10, step=1)

    extra_instr = st.text_area("💬 Extra instruction to AI (optional)",
                               placeholder="e.g. prefer olive oil brands, include at least 2 seafood items…",
                               height=80)

    st.divider()
    build_btn = st.button("🚀 Build AI List", use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN TABS
# ══════════════════════════════════════════════════════════════════════════════
tab_data, tab_list, tab_export = st.tabs(["📂 Data & Analysis", "🤖 AI Price List", "📤 Export"])

# ── TAB 1 · DATA & ANALYSIS ──────────────────────────────────────────────────
with tab_data:

    # ── Handle file analysis ───────────────────────────────────────────────
    if analyze_btn:
        files = [(f1, st.session_state.get("fu1_name", "File 1")),
                 (f2, st.session_state.get("fu2_name", "File 2"))]
        uploaded_files = [(f, f.name) for f, _ in files if f is not None]

        if not uploaded_files:
            st.warning("Please upload at least one Excel file.")
        else:
            dfs, mappings, names = [], [], []
            analysis_parts = []

            with st.spinner("AI is reading your files…"):
                for fobj, fname in uploaded_files:
                    df_raw = read_any_excel(fobj)
                    if df_raw is None:
                        st.error(f"Could not read {fname}")
                        continue

                    mapping = ai_detect_columns(df_raw, fname)
                    explanation = mapping.pop("explanation", "")
                    analysis_parts.append(f"**{fname}**: {explanation}")

                    df_norm = normalise_df(df_raw, mapping, fname)
                    dfs.append(df_norm)
                    mappings.append(mapping)
                    names.append(fname)

            if dfs:
                unified = pd.concat(dfs, ignore_index=True)
                st.session_state.dfs = dfs
                st.session_state.file_names = names
                st.session_state.mappings = mappings
                st.session_state.unified = unified
                st.session_state.ai_analysis = "\n\n".join(analysis_parts)

                with st.spinner("AI is inferring cuisines and categories…"):
                    cuisines, categories = ai_infer_cuisines_and_categories(unified)
                    st.session_state.detected_cuisines = cuisines
                    st.session_state.detected_categories = categories

                st.success(f"✅ Analysed {len(dfs)} file(s) → {len(unified)} products loaded")
                st.rerun()

    # ── Show analysis results ──────────────────────────────────────────────
    if st.session_state.unified is not None:
        unified = st.session_state.unified

        st.markdown('<div class="st">📊 Analysis Summary</div>', unsafe_allow_html=True)
        if st.session_state.ai_analysis:
            st.markdown(
                f'<div class="ai-box"><div class="ai-title">🤖 AI Column Detection</div>'
                f'{st.session_state.ai_analysis}</div>',
                unsafe_allow_html=True,
            )

        # Metrics
        c1,c2,c3,c4 = st.columns(4)
        with c1:
            st.markdown(f'<div class="card"><div class="val">{len(unified)}</div><div class="lbl">Total Products</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="card"><div class="val">{len(st.session_state.file_names)}</div><div class="lbl">Files Loaded</div></div>', unsafe_allow_html=True)
        with c3:
            n_priced = unified["price"].notna().sum()
            st.markdown(f'<div class="card"><div class="val">{n_priced}</div><div class="lbl">With Price</div></div>', unsafe_allow_html=True)
        with c4:
            n_cats = unified["category"].nunique()
            st.markdown(f'<div class="card"><div class="val">{n_cats}</div><div class="lbl">Categories</div></div>', unsafe_allow_html=True)

        # Detected metadata
        if st.session_state.detected_cuisines:
            st.markdown('<div class="st">🍽️ Detected Cuisines</div>', unsafe_allow_html=True)
            chips = " ".join(f'<span class="chip">{c}</span>' for c in st.session_state.detected_cuisines)
            st.markdown(chips, unsafe_allow_html=True)

        if st.session_state.detected_categories:
            st.markdown('<div class="st">📦 Detected Categories</div>', unsafe_allow_html=True)
            chips = " ".join(f'<span class="chip chip-g">{c}</span>' for c in st.session_state.detected_categories)
            st.markdown(chips, unsafe_allow_html=True)

        # Column mappings per file
        st.markdown('<div class="st">🗂️ Column Mapping per File</div>', unsafe_allow_html=True)
        for fname, mapping in zip(st.session_state.file_names, st.session_state.mappings):
            with st.expander(f"📄 {fname}"):
                for k, v in mapping.items():
                    if k != "explanation":
                        icon = "✅" if v else "❌"
                        st.markdown(f"{icon} **{k}** → `{v or 'not found'}`")

        # Preview table
        st.markdown('<div class="st">👁️ Unified Product Preview</div>', unsafe_allow_html=True)
        preview = unified.head(100).copy()
        preview["price"] = preview["price"].apply(lambda x: f"{x:.3f}" if pd.notna(x) else "—")
        st.dataframe(preview, use_container_width=True, height=360)

    else:
        st.info("👆 Upload 1 or 2 Excel files in the sidebar, then click **Analyse Files with AI**.")


# ── TAB 2 · AI PRICE LIST ────────────────────────────────────────────────────
with tab_list:

    if build_btn:
        if st.session_state.unified is None:
            st.warning("Analyse your files first (go to Data & Analysis tab).")
        else:
            unified = st.session_state.unified

            # Apply category filter if set
            df_filtered = unified.copy()
            if selected_cats:
                mask = df_filtered["category"].astype(str).apply(
                    lambda c: any(s.lower() in c.lower() for s in selected_cats)
                )
                df_filtered = df_filtered[mask]
                if len(df_filtered) < n_products:
                    st.warning(f"Only {len(df_filtered)} products match those categories — relaxing filter.")
                    df_filtered = unified.copy()

            with st.spinner(f"AI is curating the best {n_products} products for {selected_cuisine} cuisine…"):
                try:
                    items, reasoning = ai_build_list(
                        df_filtered, selected_cuisine, selected_cats, n_products, extra_instr
                    )
                    st.session_state.generated_list = items
                    st.session_state.ai_reasoning = reasoning
                    st.session_state.list_cuisine = selected_cuisine
                    st.success(f"✅ {len(items)} products selected for **{selected_cuisine}** cuisine!")
                    st.rerun()
                except Exception as e:
                    st.error(f"AI error: {e}")

    if st.session_state.generated_list:
        items = st.session_state.generated_list
        cuisine = st.session_state.get("list_cuisine", "")

        st.markdown(
            f'<div class="st">🤖 AI-Selected: {cuisine} Cuisine Price List</div>',
            unsafe_allow_html=True,
        )

        # AI reasoning
        if st.session_state.ai_reasoning:
            st.markdown(
                f'<div class="ai-box"><div class="ai-title">🧠 AI Reasoning</div>'
                f'{st.session_state.ai_reasoning}</div>',
                unsafe_allow_html=True,
            )

        # Totals
        total = sum(float(i["price"]) for i in items if i.get("price") not in (None, "", "nan"))
        avg   = total / len(items) if items else 0
        ca, cb, cc = st.columns(3)
        with ca:
            st.markdown(f'<div class="card"><div class="val">{len(items)}</div><div class="lbl">Products</div></div>', unsafe_allow_html=True)
        with cb:
            st.markdown(f'<div class="card"><div class="val">{total:.3f}</div><div class="lbl">Total (TND)</div></div>', unsafe_allow_html=True)
        with cc:
            st.markdown(f'<div class="card"><div class="val">{avg:.3f}</div><div class="lbl">Avg Price (TND)</div></div>', unsafe_allow_html=True)

        # Product rows
        st.markdown("")
        for i, item in enumerate(items, 1):
            price_str = f"{float(item['price']):.3f} TND" if item.get("price") not in (None, "") else "—"
            src_color = "#F97316" if item["source"] == st.session_state.file_names[0] else "#16A34A"
            st.markdown(
                f"""<div class="prow">
                  <span style="color:#78716C;font-size:.8rem;min-width:1.5rem;font-weight:700">{i}</span>
                  <span class="pname">{item['name']}</span>
                  <span class="psup">{item.get('supplier','') or '—'}</span>
                  <span style="color:#A8A29E;font-size:.8rem">{item['cond'] or '—'}</span>
                  <span class="pprice">{price_str}</span>
                  <span class="pcat">{item['cat'] or '—'}</span>
                  <span style="font-size:.7rem;font-weight:700;color:{src_color}">{item['source']}</span>
                </div>""",
                unsafe_allow_html=True,
            )

        st.divider()
        if st.button("🔄 Regenerate with AI"):
            st.session_state.generated_list = None
            st.session_state.ai_reasoning = ""
            st.rerun()

    else:
        st.info("Set your cuisine and options in the sidebar, then click **🚀 Build AI List**.")


# ── TAB 3 · EXPORT ───────────────────────────────────────────────────────────
with tab_export:

    if not st.session_state.generated_list:
        st.info("Generate a price list first (AI Price List tab).")
    else:
        items   = st.session_state.generated_list
        cuisine = st.session_state.get("list_cuisine", "List")
        title   = f"Foodomarket – {cuisine} Price List {datetime.date.today().strftime('%Y-%m-%d')}"

        st.markdown('<div class="st">📤 Export Your Price List</div>', unsafe_allow_html=True)
        st.markdown(f"**{len(items)} products** · **{cuisine} cuisine**")

        c1, c2 = st.columns(2)
        with c1:
            xlsx_bytes = export_excel(items, title)
            st.download_button(
                "⬇️ Download Excel (.xlsx)",
                data=xlsx_bytes,
                file_name=f"{title.replace(' ','_').replace('–','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with c2:
            pdf_bytes = export_pdf(items, title, cuisine)
            st.download_button(
                "⬇️ Download PDF (.pdf)",
                data=pdf_bytes,
                file_name=f"{title.replace(' ','_').replace('–','')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        st.divider()
        st.markdown('<div class="st">📋 Preview Table</div>', unsafe_allow_html=True)
        preview_df = pd.DataFrame([{
            "#": i+1,
            "Product": it["name"],
            "Conditioning": it["cond"] or "—",
            "Price (TND)": f"{float(it['price']):.3f}" if it.get("price") not in (None,"") else "—",
            "Category": it["cat"] or "—",
            "Supplier": it.get("supplier") or "—",
            "Source File": it["source"],
        } for i, it in enumerate(items)])
        st.dataframe(preview_df, use_container_width=True, hide_index=True)
